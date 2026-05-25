#!/usr/bin/env python3
"""Excel工程量清单读取与清洗模块"""

import os
import re
import hashlib
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl

import config


@dataclass
class CleanedItem:
    """清洗后的一条记录"""
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0
    import_batch: str = ""
    import_time: str = ""

    # 业务字段
    项目名称: str = ""
    项目编号: str = ""
    项目特征描述: str = ""
    计量单位: str = ""
    工程数量: Optional[float] = None
    不含税单价: Optional[float] = None
    汇总合价: Optional[float] = None
    合同名称: str = ""

    # 定价信息
    定价信息: str = ""

    # 分类字段
    页签: str = ""
    层级路径: str = ""
    二级页签: str = ""  # 兼容旧脚本：仅保留原二级标题，不再作为导出字段
    科目: str = ""      # 兼容旧脚本：仅保留科目路径，不再作为导出字段

    # 异常标记
    异常标记: str = ""
    单价为空: bool = False
    工程量为空: bool = False
    合价异常: bool = False
    疑似有效行: bool = False

    def source_hash(self) -> str:
        """基于业务字段生成去重hash"""
        raw = "|".join([
            str(getattr(self, k, "")) for k in config.DEDUP_KEYS
        ])
        return hashlib.md5(raw.encode()).hexdigest()


@dataclass
class AnomalyItem:
    """异常记录"""
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0
    异常类型: str = ""
    异常描述: str = ""
    项目名称: str = ""
    项目编号: str = ""


# ----- 合并单元格处理 -----

def get_cell_value(ws, row: int, col: int):
    """读取单元格值，处理合并单元格"""
    val = ws.cell(row, col).value
    if val is not None:
        return val
    for merge_range in ws.merged_cells.ranges:
        if (merge_range.min_col <= col <= merge_range.max_col
                and merge_range.min_row <= row <= merge_range.max_row):
            top_val = ws.cell(merge_range.min_row, merge_range.min_col).value
            return top_val
    return None


# ----- 表头锚点定位 -----

def find_header_anchor(ws) -> Optional[Tuple[int, int]]:
    """找到表头锚点行和项目编号列。返回 (header_row, col_of_项目编号)"""
    def has_name_header(row: int, anchor_col: int) -> bool:
        for c2 in range(1, min(ws.max_column + 1, 20)):
            v2 = str(get_cell_value(ws, row, c2) or "").strip()
            if "项目名称" in v2:
                return True
        typo_col = anchor_col + 1
        typo = str(get_cell_value(ws, row, typo_col) or "").strip()
        return typo == "项"

    for row in range(1, min(ws.max_row + 1, 30)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = str(get_cell_value(ws, row, col) or "").strip()
            if "项目编号" in val:
                if has_name_header(row, col):
                    return row, col
    # 回退：找 项目编码
    for row in range(1, min(ws.max_row + 1, 30)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = str(get_cell_value(ws, row, col) or "").strip()
            if "项目编码" in val:
                if has_name_header(row, col):
                    return row, col
    # 回退：部分精装清单第一列为“序号”，没有“项目编号”
    for row in range(1, min(ws.max_row + 1, 30)):
        for col in range(1, min(ws.max_column + 1, 20)):
            val = str(get_cell_value(ws, row, col) or "").strip()
            if val == "序号":
                if has_name_header(row, col):
                    return row, col
    return None


# ----- 定价模式检测 + 列扫描 -----

def detect_pricing_mode(ws, header_row: int, anchor_col: int) -> str:
    """检测定价模式 A普通 / B基准价。返回 'A' 或 'B'
    支持：
    - 标准B型：表头含 基准价+浮率+投标价
    - 下浮后B型：上一行有"下浮后"标记，表头有综合单价（原价+折后价）
    """
    limit_row = min(header_row + 4, ws.max_row + 1)
    limit_col = min(anchor_col + 40, ws.max_column + 1)
    block_has_base = False
    block_has_rate = False
    block_has_bid = False
    for row in range(header_row, limit_row):
        has_base = False
        has_rate = False
        has_bid = False
        for col in range(anchor_col, limit_col):
            val = str(get_cell_value(ws, row, col) or "").strip()
            if "基准价" in val or "基准单价" in val:
                has_base = True
                block_has_base = True
            if "浮率" in val:
                has_rate = True
                block_has_rate = True
            if "投标价" in val:
                has_bid = True
                block_has_bid = True
        if has_base and has_rate and has_bid:
            return "B"
    if block_has_base and block_has_rate and block_has_bid:
        return "B"

    # 下浮后模式检测：上一行有"下浮后"标记
    if header_row > 1:
        has_discount = False
        for col in range(anchor_col, limit_col):
            v = str(get_cell_value(ws, header_row - 1, col) or "").strip()
            if "下浮后" in v:
                has_discount = True
                break
        if has_discount:
            return "B"  # 下浮后模式视为B型（基准价为原价，投标价为折后价）

    return "A"


def scan_column_map(ws, header_row: int, anchor_col: int) -> dict:
    """扫描表头行，确定各列的业务用途。
    以主表头行（header_row）为准，子表头行仅补充。
    支持下浮后模式：header_row-1 行标记"下浮后"的列为投标价列。
    """
    col_map = {}

    limit_col = min(anchor_col + 40, ws.max_column + 1)

    # 第0遍：扫描 header_row-1 行，找出标记为"下浮后"的列
    discounted_cols = set()
    if header_row > 1:
        # 检测上一行是否有"下浮后"标记
        has_discount_label = False
        for c in range(anchor_col, limit_col):
            v = str(get_cell_value(ws, header_row - 1, c) or "").strip()
            if "下浮后" in v:
                has_discount_label = True
                break

        # 如果有下浮后标记，在 header_row 中找出定价列的分组
        # 定价列通常分两组：C7-C9（原价）和 C12-C14（折后价）
        # 标记第二组为折扣列
        if has_discount_label:
            pricing_cols = []
            for c in range(anchor_col, limit_col):
                v = str(ws.cell(header_row, c).value or "").strip()
                if "综合单价" in v and "不含" in v:
                    pricing_cols.append(c)
                elif "不含" in v and "单价" in v:
                    pricing_cols.append(c)

            # 找列间隙：如果两列之间差>1，说明是新组的开始
            if pricing_cols:
                groups = [[pricing_cols[0]]]
                for col in pricing_cols[1:]:
                    if col - groups[-1][-1] > 1:
                        groups.append([col])
                    else:
                        groups[-1].append(col)

                # 如果有多组，最后一组就是折扣列
                if len(groups) > 1:
                    for col in groups[-1]:
                        discounted_cols.add(col)

    # 第1遍：扫描主表头行（用直接单元格值，避免合并单元格覆盖）
    for c in range(anchor_col, limit_col):
        v = str(ws.cell(header_row, c).value or "").strip()
        if not v:
            continue
        offset = c - anchor_col

        if "汇总合价" in v:
            col_map["汇总合价"] = offset  # 覆盖式赋值，让最后的"汇总合价"胜出（B型两列合价时取靠后的投标价合价）
        elif ("合价" in v or "金额" in v) and "人工" not in v and "材料" not in v:
            if "汇总合价" not in col_map:
                col_map["汇总合价"] = offset

        if "投标价" in v:
            col_map["投标价"] = offset

        if "主材价" in v:
            col_map["主材价"] = offset

        # 综合单价/不含税单价处理（支持下浮后模式）
        if "综合单价" in v and "不含" in v:
            if c in discounted_cols:
                # 下浮后综合单价 → 映射为投标价（覆盖）
                col_map["投标价"] = offset
            else:
                # 正常综合单价 → 映射为单价和基准价
                col_map["单价"] = offset
                col_map["基准价"] = offset
        elif "不含" in v and "单价" in v:
            if c in discounted_cols:
                col_map["投标价"] = offset
            else:
                col_map["单价"] = offset

        if ("基准价" in v or "基准单价" in v) and ("不含税" in v or "不含增值税" in v):
            col_map["基准价"] = offset

        if "浮率" in v:
            col_map["上下浮率"] = offset

        if "汇总工程量" in v:
            col_map["汇总工程量"] = offset

        if "工程数量" in v and "汇总" not in v:
            if "工程数量" not in col_map:
                col_map["工程数量"] = offset

    # 第2遍：主表头行没找到的列，用合并单元格回退的get_cell_value再查一次
    for c in range(anchor_col, limit_col):
        v = str(ws.cell(header_row, c).value or "").strip()
        if v:
            continue  # 已有直接值，跳过
        mv = str(get_cell_value(ws, header_row, c) or "").strip()
        if not mv:
            continue
        offset = c - anchor_col
        
        if "投标价" in mv and "投标价" not in col_map:
            col_map["投标价"] = offset
        if "主材价" in mv and "主材价" not in col_map:
            col_map["主材价"] = offset
        if "汇总合价" in mv and "汇总合价" not in col_map:
            col_map["汇总合价"] = offset
        if "汇总工程量" in mv and "汇总工程量" not in col_map:
            col_map["汇总工程量"] = offset

    # 第2.5遍：支持“下浮后综合单价”分组。
    # 上一行是分组标题，当前行仍是“不含增值税综合单价（元）”。
    if header_row > 1:
        for c in range(anchor_col, limit_col):
            prev = str(get_cell_value(ws, header_row - 1, c) or "").strip()
            current = str(ws.cell(header_row, c).value or "").strip()
            if "下浮后" in prev and "综合单价" in current and "单价分析" not in prev:
                offset = c - anchor_col
                col_map["单价"] = offset
                col_map["投标价"] = offset

    # 第3遍：如果关键列还没找到，查子表头行（也采用覆盖式，让最后出现的合价列胜出）
    if "单价" not in col_map or "汇总合价" not in col_map:
        limit_sub = min(header_row + 4, ws.max_row + 1)
        for r in range(header_row + 1, limit_sub):
            for c in range(anchor_col, limit_col):
                v = str(get_cell_value(ws, r, c) or "").strip()
                if not v:
                    continue
                offset = c - anchor_col

                if "单价" not in col_map:
                    if "综合单价" in v or ("不含" in v and "税" in v):
                        col_map["单价"] = offset
                if "汇总合价" in v or v == "合价（元）":
                    col_map["汇总合价"] = offset  # 覆盖式，取靠后的投标价合价
                if "投标价" not in col_map and "投标价" in v:
                    col_map["投标价"] = offset
                if "主材价" not in col_map and "主材价" in v:
                    col_map["主材价"] = offset
                if "汇总工程量" not in col_map and "汇总工程量" in v:
                    col_map["汇总工程量"] = offset

    # 第4遍：处理两层表头。
    # 例如顶层“投标价（不含增值税）”横跨“综合单价/汇总合价”两列，
    # 必须用子表头把投标价单价和投标价合价拆开，否则会误取合价列为单价。
    limit_sub = min(header_row + 4, ws.max_row + 1)
    for r in range(header_row + 1, limit_sub):
        for c in range(anchor_col, limit_col):
            top = str(get_cell_value(ws, header_row, c) or "").strip()
            sub = str(ws.cell(r, c).value or "").strip()
            if not top or not sub:
                continue
            offset = c - anchor_col
            if "基准价" in top or "基准单价" in top:
                if "综合单价" in sub:
                    col_map["基准价"] = offset
                    col_map.setdefault("单价", offset)
                elif "汇总合价" in sub:
                    col_map.setdefault("基准合价", offset)
            if "投标价" in top:
                if "综合单价" in sub:
                    col_map["投标价"] = offset
                elif "汇总合价" in sub:
                    col_map["汇总合价"] = offset

    # 单层B型表头可能只有“投标价-不含增值税”合并跨两列：
    # 第一列是投标单价，后一列是投标合价。
    if "投标价" in col_map and "汇总合价" not in col_map:
        candidate = anchor_col + col_map["投标价"] + 1
        if candidate < limit_col:
            top = str(get_cell_value(ws, header_row, candidate) or "").strip()
            direct = str(ws.cell(header_row, candidate).value or "").strip()
            if not direct and "投标价" in top:
                col_map["汇总合价"] = candidate - anchor_col

    return col_map


# ----- 行类型判断 -----

def is_title_row(row_text: str) -> bool:
    """是否为表头行"""
    text = row_text.strip()
    for kw in config.TITLE_KEYWORDS:
        if kw in text:
            return True
    return False


def should_skip_row(row_text: str) -> bool:
    """是否为汇总/说明行"""
    text = row_text.strip()
    for kw in config.SKIP_KEYWORDS:
        if kw in text:
            return True
    return False


# ----- 数值解析 -----

def parse_number(val) -> Optional[float]:
    """尝试将值转换为float。空值/'-'/None 返回 None"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip()
    if text == "" or text == "-" or text == "\u2014":
        return None
    text = text.replace(",", "").replace("\uff0c", "").replace(" ", "")
    try:
        return float(text)
    except:
        return None


def format_float_rate(float_rate: float) -> str:
    """Format a source float-rate number for display in pricing info."""
    return f"{float_rate * 100:.2f}".rstrip("0").rstrip(".") + "%"


def sum_numeric_cells(ws, row: int, start_col: int, end_col: int) -> Optional[float]:
    """汇总一段单元格里的数值。没有任何数值时返回 None。"""
    total = 0.0
    found = False
    for col in range(start_col, end_col + 1):
        value = parse_number(get_cell_value(ws, row, col))
        if value is not None:
            total += value
            found = True
    return total if found else None


def price_from_supply_install(ws, header_row: int, row: int, anchor_col: int, total_offset: int) -> Optional[float]:
    """总综合单价为空时，尝试用同组供应综合单价+安装综合单价补足。"""
    total_col = anchor_col + total_offset
    if total_col - 2 < anchor_col:
        return None
    h1 = str(ws.cell(header_row, total_col - 2).value or "")
    h2 = str(ws.cell(header_row, total_col - 1).value or "")
    if "供应" not in h1 or "安装" not in h2:
        return None
    p1 = parse_number(get_cell_value(ws, row, total_col - 2))
    p2 = parse_number(get_cell_value(ws, row, total_col - 1))
    if p1 is None and p2 is None:
        return None
    return (p1 or 0.0) + (p2 or 0.0)


# ----- 辅助 -----

SKIP_SHEET_KEYWORDS = [
    "封皮", "编制说明", "汇总表", "综合单价分析",
    "单价分析表", "单价分析", "开办费", "附表", "计日工", "照管费",
    "甲供材", "暂列", "备件品", "面积", "WpsReserved",
    "招标清单列表",
]


def should_process_sheet(sheet_name: str) -> bool:
    """判断sheet是否应该被处理"""
    name = sheet_name.strip()
    for kw in SKIP_SHEET_KEYWORDS:
        if kw in name:
            return False
    # 主清单sheet在不同合同中命名不一致：
    # 实体工程量清单、实体工程清单、实体清单、2.2-精装工程量清单等都应处理。
    if "汇总" in name:
        return False
    if (
        "实体工程量清单" in name
        or "实体工程清单" in name
        or "实体清单" in name
        or "工程量清单" in name
    ):
        return True
    return False


def extract_专业名称(sheet_name: str) -> str:
    """从sheet名提取专业名称"""
    name = re.sub(r"^\s*\d+(?:\.\d+)*\s*[-、.：:]*\s*", "", sheet_name.strip())
    name = re.sub(r"^(通用工程类|实体工程类)\s*[-、.：:]*\s*", "", name)
    for token in ["实体工程量清单", "实体工程清单", "工程量清单", "实体清单"]:
        name = name.replace(token, "")
    name = name.strip(" -_—、.：:")
    return name or sheet_name.strip()


def join_path(*parts: str) -> str:
    """用 / 合并层级，自动忽略空值并去重相邻重复段。"""
    result = []
    for part in parts:
        if not part:
            continue
        for piece in str(part).replace("|", "/").split("/"):
            piece = piece.strip()
            if piece and (not result or result[-1] != piece):
                result.append(piece)
    return "/".join(result)


def extract_contract_name(file_name: str) -> str:
    """从文件名提取合同名称"""
    name = os.path.splitext(file_name)[0]
    name = name.lstrip("~$")
    name = re.sub(r"^\s*\d+\s*[、，,\-_—:：;；]\s*", "", name)
    name = re.sub(r"^\s*\d+\s*[.．。]\s+(?=\S)", "", name)
    name = name.lstrip(" \t\r\n-_—、，,.:：;；。．")
    if len(name) > 80:
        name = name[:80]
    return name.strip()


# ----- 主清洗函数 -----

def clean_excel(file_path: str) -> Tuple[List[CleanedItem], List[AnomalyItem]]:
    """清洗一个Excel文件，返回 (items, anomalies)"""
    items: List[CleanedItem] = []
    anomalies: List[AnomalyItem] = []

    file_name = os.path.basename(file_path)
    if file_name.startswith("~$"):
        return items, anomalies  # 跳过临时文件

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        anomalies.append(AnomalyItem(
            source_file=file_name, 异常类型="文件打开失败",
            异常描述=str(e)
        ))
        return items, anomalies

    contract_name = extract_contract_name(file_name)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        if not should_process_sheet(sheet_name):
            continue

        if ws.max_row < 5 or ws.max_column < 5:
            continue

        anchor = find_header_anchor(ws)
        if anchor is None:
            anomalies.append(AnomalyItem(
                source_file=file_name, source_sheet=sheet_name,
                异常类型="无法定位表头",
                异常描述="找不到项目编号+项目名称所在行"
            ))
            continue

        header_row, col_offset = anchor
        mode = detect_pricing_mode(ws, header_row, col_offset)
        col_map = scan_column_map(ws, header_row, col_offset)

        # 检测是否"下浮后"模式（有折后价但无折后合价列，跳过合价校验）
        no_sum_verify = False
        for r in range(max(1, header_row - 2), header_row):
            for c in range(1, min(ws.max_column + 1, 20)):
                v = str(ws.cell(r, c).value or "").strip()
                if "下浮后" in v or "折后" in v:
                    no_sum_verify = True
                    break
            if no_sum_verify:
                break

        # 分类继承缓存
        section_categories = {}
        page_tab = extract_专业名称(sheet_name)
        sub_page_tab = ""

        data_start = header_row + 3  # 跳过合并表头

        for row in range(data_start, ws.max_row + 1):
            # 使用动态列映射读取
            try:
                # 定价模式决定列位置
                if mode == "B":
                    # B型: 用检测到的列偏移
                    bid_off = col_map.get("投标价", 9)
                    sum_off = col_map.get("汇总合价", 10)
                    qty_off = col_map.get("汇总工程量", 5)  # 优先汇总工程量，回退到工程数量
                    sum_qty_off = col_map.get("汇总工程量")
                    base_off = col_map.get("基准价")
                    rate_off = col_map.get("上下浮率")
                    material_off = col_map.get("主材价")

                    pid = get_cell_value(ws, row, col_offset + 0)
                    pname = get_cell_value(ws, row, col_offset + 1)
                    pdesc = get_cell_value(ws, row, col_offset + 2)
                    cat = str(get_cell_value(ws, row, col_offset + 3) or "").strip()
                    unit = get_cell_value(ws, row, col_offset + 4)
                    qty = get_cell_value(ws, row, col_offset + qty_off)
                    sum_qty = get_cell_value(ws, row, col_offset + sum_qty_off) if sum_qty_off is not None else None
                    bid_price = get_cell_value(ws, row, col_offset + bid_off)
                    sum_price = get_cell_value(ws, row, col_offset + sum_off)
                    unit_price = parse_number(bid_price)
                    if unit_price is None:
                        unit_price = price_from_supply_install(ws, header_row, row, col_offset, bid_off)

                    # 提取基准价和上下浮率（用于定价信息）
                    base_price = parse_number(get_cell_value(ws, row, col_offset + base_off)) if base_off is not None else None
                    float_rate = parse_number(get_cell_value(ws, row, col_offset + rate_off)) if rate_off is not None else None
                    material_price = parse_number(get_cell_value(ws, row, col_offset + material_off)) if material_off is not None else None

                    # 生成定价信息 + 浮率验证
                    pricing_info = ""
                    if base_price is not None and float_rate is not None:
                        pricing_info = f"基准价:{base_price}, 上下浮率:{format_float_rate(float_rate)}"
                        # 浮率验证：区分两种语义
                        # |浮率| > 0.5 → 乘数模式（0.988=98.8%，直接乘）
                        # |浮率| <= 0.5 → 幅度模式（-0.083=-8.3%，基准价×(1+浮率)）
                        if unit_price is not None:
                            if abs(float_rate) > 0.5:
                                expected = round(base_price * float_rate, 2)
                                formula = f"基准价{base_price}×{float_rate}"
                                if material_price is not None:
                                    adjusted_expected = round(material_price + (base_price - material_price) * float_rate, 2)
                                    if abs(unit_price - adjusted_expected) <= 0.05:
                                        expected = adjusted_expected
                                        formula = f"主材价{material_price}+(基准价{base_price}-主材价{material_price})×{float_rate}"
                            else:
                                expected = round(base_price * (1 + float_rate), 2)
                                formula = f"基准价{base_price}×(1+{float_rate})"
                            if abs(unit_price - expected) > 0.05:
                                pid_preview = str(pid or "").strip()
                                print("\n浮率异常！导入已停止")
                                print(f"  投标价={unit_price} ≠ 计算值={expected}")
                                print(f"  公式: {formula}")
                                print(f"  文件: {file_name} | Sheet: {sheet_name} | 行: {row}")
                                print(f"  项目编号: {pid_preview}")
                                print(f"  请检查数据源，修复后重新运行。\n")
                                raise SystemExit(1)  # 停止导入
                    elif base_price is not None and unit_price is not None:
                        # 下浮后模式：无浮率列，从比值计算
                        if base_price:
                            implied_rate = round(unit_price / base_price, 6)
                            pricing_info = f"基准价:{base_price}, 上下浮率:{format_float_rate(implied_rate)}"
                        else:
                            pricing_info = f"基准价:{base_price}"
                    elif base_price is not None:
                        pricing_info = f"基准价:{base_price}"
                    else:
                        pricing_info = "自主报价"
                else:
                    # A型: 用检测到的列偏移
                    price_off = col_map.get("单价", 6)
                    sum_off = col_map.get("汇总合价", 7)
                    qty_off = col_map.get("汇总工程量", col_map.get("工程数量", 5))
                    sum_qty_off = col_map.get("汇总工程量")

                    pid = get_cell_value(ws, row, col_offset + 0)
                    pname = get_cell_value(ws, row, col_offset + 1)
                    pdesc = get_cell_value(ws, row, col_offset + 2)
                    cat = str(get_cell_value(ws, row, col_offset + 3) or "").strip()
                    unit = get_cell_value(ws, row, col_offset + 4)
                    qty = get_cell_value(ws, row, col_offset + qty_off)
                    sum_qty = get_cell_value(ws, row, col_offset + sum_qty_off) if sum_qty_off is not None else None
                    raw_price = get_cell_value(ws, row, col_offset + price_off)
                    sum_price = get_cell_value(ws, row, col_offset + sum_off)
                    unit_price = parse_number(raw_price)
                    if unit_price is None:
                        unit_price = price_from_supply_install(ws, header_row, row, col_offset, price_off)

                    # A型：自主报价
                    pricing_info = "自主报价"
            except Exception as e:
                print(f"  读取失败 {row}: {e}")
                continue

            pid_s = str(pid or "").strip()
            pname_s = str(pname or "").strip()
            pdesc_s = str(pdesc or "").strip()
            unit_s = str(unit or "").strip()
            if not unit_s and pname_s == "马桶（甲供）":
                unit_s = "套"
            qty_f = parse_number(qty)
            sum_qty_f = parse_number(sum_qty) if sum_qty is not None else None
            if qty_f is None and sum_qty_f is None and sum_qty_off is not None and sum_qty_off > 5:
                qty_f = sum_numeric_cells(ws, row, col_offset + 5, col_offset + sum_qty_off - 1)
            sum_price_f = parse_number(sum_price)
            
            # A型：如果汇总合价值明显不合理，忽略
            if mode == "A" and sum_price_f is not None and qty_f and unit_price:
                if sum_price_f > qty_f * unit_price * 10:
                    sum_price_f = None

            # 跳过完全空行
            if not pid_s and not pname_s and not pdesc_s:
                continue
            # 跳过只有编号/补项号、没有名称和单位的占位行。
            if pid_s and not pname_s and not pdesc_s and not unit_s:
                continue

            row_text = " ".join([pid_s, pname_s, pdesc_s])

            # 跳过表头
            if is_title_row(row_text):
                continue

            # 跳过汇总
            if should_skip_row(row_text):
                continue

            # === 分类继承 ===
            if cat == "科目":
                section_categories[pid_s] = pname_s
                # 清理子级
                for k in list(section_categories.keys()):
                    if k.startswith(pid_s + ".") or (len(k) > len(pid_s) and k.startswith(pid_s[:len(pid_s)])):
                        pass  # 保留父级
                continue

            if cat in ("页签", "二级页签"):
                sub_page_tab = pname_s
                continue

            if cat == "三级页签":
                sub_page_tab = join_path(sub_page_tab, pname_s)
                continue

            # === 判断是否有效清单行 ===
            is_list = (cat == "清单项")
            has_unit = bool(unit_s)
            has_qty = (qty_f is not None)
            has_price = (unit_price is not None)

            # 金额汇总行（无单位无数量）跳过
            if not is_list and not has_unit and not has_qty:
                if not pid_s and not pname_s:
                    continue
                if pid_s and pname_s and not has_unit and not has_qty:
                    continue  # 层次行
            
            # 无编号、无单价且有名称 → 视为同级标题，切换当前二级标题。
            # 这类行经常没有显式“二级页签”类别，不能追加到上一标题后面。
            if not pid_s and not has_price and pname_s and pdesc_s == "":
                sub_page_tab = pname_s
                continue
            
            # 无类别但有数量或合价且无单价 → 汇总行，跳过
            if not cat and not has_price and (has_qty or sum_price_f is not None):
                continue
            
            # 清单项但无单位无单价 → 实为同级标题，切换当前二级标题。
            if is_list and not unit_s and not has_price:
                sub_page_tab = pname_s
                continue

            # 部分清单把层级标题误标为“清单项”，常见特征是单位为“项”，
            # 名称和特征描述相同，且数量/单价/合价均为空。
            if (
                is_list
                and unit_s == "项"
                and qty_f is None
                and unit_price is None
                and sum_price_f is None
                and pname_s
            ):
                sub_page_tab = pname_s
                continue

            # 清单项中无单位、无数量且金额为 0/空的行，多为专业占位或分组行。
            if (
                is_list
                and not unit_s
                and (qty_f in (None, 0.0))
                and (unit_price in (None, 0.0))
                and (sum_price_f in (None, 0.0))
                and pname_s
            ):
                sub_page_tab = pname_s
                continue

            # 非清单项、无单位且金额数量均为0的行，通常是占位或无效行。
            if not is_list and not unit_s and (qty_f in (None, 0.0)) and (unit_price in (None, 0.0)) and (sum_price_f in (None, 0.0)):
                continue

            # 疑似有效行：有名称且有单位或数量
            is_suspect = False
            if not is_list and (has_unit or has_qty) and pname_s:
                is_suspect = True

            # === 构建记录 ===
            subject_path = "|".join([v for k, v in sorted(section_categories.items(), key=lambda x: len(x[0]))])
            layer_path = join_path(subject_path, sub_page_tab) or page_tab

            item = CleanedItem(
                source_file=file_name,
                source_sheet=sheet_name,
                source_row=row,
                项目名称=pname_s,
                项目编号=pid_s,
                项目特征描述=pdesc_s,
                计量单位=unit_s,
                工程数量=qty_f,
                不含税单价=unit_price,
                汇总合价=sum_price_f,
                合同名称=contract_name,
                页签=page_tab,
                层级路径=layer_path,
                二级页签=sub_page_tab,
                科目=subject_path,
                单价为空=(unit_price is None and is_list),
                工程量为空=(qty_f is None and is_list),
                疑似有效行=is_suspect,
                定价信息=pricing_info,
            )

            # 合价校验 (下浮后模式跳过，因无折后合价列)
            if not no_sum_verify:
                verify_qty = sum_qty_f or qty_f
                if verify_qty is not None and unit_price is not None and sum_price_f is not None:
                    expected = round(verify_qty * unit_price, 2)
                    diff = abs(sum_price_f - expected)
                    if diff > 0.05:
                        item.合价异常 = True
                        item.异常标记 = "合价异常"

            # 异常标记
            notes = []
            if item.单价为空:
                notes.append("单价为空")
            if item.工程量为空:
                notes.append("工程量为空")
            if item.合价异常:
                notes.append("合价异常")
            if item.疑似有效行:
                notes.append("疑似有效行")
            if notes:
                item.异常标记 = "|".join(notes)

            items.append(item)

    # 用同文件内同名项目的唯一单位补齐源表偶发漏填单位。
    units_by_name: Dict[str, set] = {}
    for item in items:
        if item.项目名称 and item.计量单位:
            units_by_name.setdefault(item.项目名称, set()).add(item.计量单位)
    for item in items:
        if item.项目名称 and not item.计量单位:
            units = units_by_name.get(item.项目名称, set())
            if len(units) == 1:
                item.计量单位 = next(iter(units))

    wb.close()
    return items, anomalies

